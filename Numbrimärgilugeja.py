################################################
# Programmeerimine I
# 2024/2025 sügissemester
#
# Projekt
# Teema: 
# Auto numbrimärgi lugemine ja siis andmebaasist vastava numbrimärgiga andmete võtmine ja muutmine
#
# Autorid:
# Andreas Jürgenson
# Richard Jaarman
# mõningane eeskuju:
#
# Lisakommentaar (nt käivitusjuhend):
# Tuleb tõmmata pytesseract nii pythonis kui ka arvutisse (https://github.com/UB-Mannheim/tesseract/wiki)(otsida ka tesseracti asukoht)
# pip install Openpyxl,Pandas, customtkinter
# Pildifailid ei tohi sisaldada öäüõ
##################################################
import customtkinter
import tkinter
from ctypes import windll
from PIL import Image,ImageTk
from pytesseract import pytesseract #Teksti saamiseks
import cv2 #pildi töötluse jaoks
import re #tekstimustrite leidmiseks
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook,Workbook

#Nupufunktsioonid
# def pildista():
#     print("Pildistan!")
def valipilt():
    global pilttt, img_label,backup
    pilttt = customtkinter.filedialog.askopenfilename() # Valib faili
    if pilttt=="": # kui vajutada loobu
        if backup!=None:
            pilttt=backup
        else:
            return
    else:
        backup=pilttt
    if pilttt:
        fail=pilttt.split("/")
        nupp1.configure(text=fail[-1])
            # avab pildi
        img=Image.open(pilttt)
        img.thumbnail((300,300))
        img_tk = ImageTk.PhotoImage(img)
        if img_label is None:
            img_label=tkinter.Label(master=app, image=img_tk, bg="white")
            img_label.image = img_tk
            img_label.place(relx=0.24, rely=0.5, anchor=tkinter.CENTER)
        else:
            img_label.configure(image=img_tk)
            img_label.image = img_tk
def valiexcel():
    global excel
    excel = customtkinter.filedialog.askopenfilename() #valib exceli faili
    fail=excel.split("/")
    nupp2.configure(text=fail[-1])
def kas_on_numbrim2rk(numbrim2rgitekst):
    if len(numbrim2rgitekst) == 6 and numbrim2rgitekst[:3].isdigit() and numbrim2rgitekst[3:].isalpha():
        return True
    return False
def analuus():
    global pilttt
    global puhas_tekst
    konsool.configure(state="normal")
    konsool.delete("0.0","end")
    puhas_tekst=None
    pildi_asukoht=pilttt
    if len(nrm2rk.get())!=0:
        puhas_tekst=nrm2rk.get()
        puhas_tekst=puhas_tekst.replace(" ","").upper()
        tekstvar.set(f"Numbrimärk on {puhas_tekst}")
        konsool.insert("end","Sisesta liiva kogus")
        konsool.configure(state="disabled")
        return 
    if pilttt is None:
        konsool.insert("end","Oled jätnud pildi lisamata")
        konsool.configure(state="disabled")
        return
    try:
        sobiliknmbrm2rk = False
        pilt = cv2.imread(pildi_asukoht)
        if pilt is None:
            konsool.insert("end","Ei leia pilti")
            konsool.configure(state="disabled")
            return
        hall = cv2.cvtColor(pilt, cv2.COLOR_BGR2GRAY) # muudab värvipildi halliks, kergem keskenduda ainult heledatele asjadele
        hall = cv2.bilateralFilter(hall, 15, 75, 75) # müra vähendamiseks, numbrid näitavad filtri suurust ja intensiivsust
        hall = cv2.convertScaleAbs(hall, alpha=1.8, beta=30) #kontrasti ja heleduse muutmiseks
        ääred = cv2.Canny(hall, 50, 150) # pildi äärte leidmiseks
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3)) # aitab ääri ühendada kontuuridel
        ääred = cv2.dilate(ääred, kernel, iterations=1)
        kontuurid, _ = cv2.findContours(ääred.copy(), cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE) #kujundite äärte leidmiseks, lisaks koopia, et originaalne ei muutuks, leiab kõik kontuurid ja siis võtab ebavajalikud kontuurid ära
        cv2.imwrite("hall_pilt.jpg", ääred)
        
        kontuur_img = pilt.copy() #tekitab originaal pildile kontuurid
        cv2.drawContours(kontuur_img, kontuurid, -1, (0, 255, 0), 2)
        cv2.imwrite("kontuurid.jpg", kontuur_img)

        kontuurid = sorted(kontuurid, key=cv2.contourArea, reverse=True)[:30] # 10 suurimat kontuuri kahanevas järjekorras

        numbrimärk = None
        for kontuur in kontuurid:
            ümbermõõt = cv2.arcLength(kontuur, True) #Arvutab kontuuriümbermõõdu, True ütleb, et kontuur on suletud
            umbkaudne = cv2.approxPolyDP(kontuur, 0.01 * ümbermõõt, True) # leiab kontuuri kuju
            if 4 <= len(umbkaudne) <= 6:
                x, y, w, h = cv2.boundingRect(umbkaudne)
                aspect_ratio = w / float(h)
                if 2 < aspect_ratio < 5:  # vaatab kas proportsioonid vastavad tavalisele numbrimärgi proportsioonidele
                    numbrimärk = umbkaudne
                    break
        if numbrimärk is None:
            # Algsele pildile tagasi, proovib lisaks teist meetodit
            hall = cv2.cvtColor(pilt, cv2.COLOR_BGR2GRAY)
            
            # Teine pilditöötlus meetod, mis muudab pilti teiste filtritega, kui esimese korraga numbrimärki ei leia
            hall = cv2.GaussianBlur(hall, (5, 5), 0)
            hall = cv2.adaptiveThreshold(hall, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 11, 2)
            
            #uuesti pildi ääred ja kontuuride kujundid
            ääred = cv2.Canny(hall, 50, 150)
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
            ääred = cv2.dilate(ääred, kernel, iterations=1)
            kontuurid, _ = cv2.findContours(ääred.copy(), cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
            kontuurid = sorted(kontuurid, key=cv2.contourArea, reverse=True)[:30]

            # kontrollib kas nüüd on numbrimärki leitud
            for kontuur in kontuurid:
                ümbermõõt = cv2.arcLength(kontuur, True)
                umbkaudne = cv2.approxPolyDP(kontuur, 0.01 * ümbermõõt, True)
                if 4 <= len(umbkaudne) <= 6:
                    x, y, w, h = cv2.boundingRect(umbkaudne)
                    aspect_ratio = w / float(h)
                    if 2 < aspect_ratio < 5:
                        numbrimärk = umbkaudne
                        break
        if numbrimärk is not None:
            x, y, w, h = cv2.boundingRect(numbrimärk) # numbrimärgi kordinaadid ja mõõtmed
            numbrimärk_pilt = hall[y:y + h, x:x + w]
            numbrimärk_pilt = cv2.resize(numbrimärk_pilt, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
            numbrimärk_pilt = cv2.GaussianBlur(numbrimärk_pilt, (5, 5), 0)
            cv2.imwrite("nmbrim2rk.jpg", numbrimärk_pilt)
            numbrimärigi_tähed = pytesseract.image_to_string(numbrimärk_pilt, config='--psm 7') #juhendab otsima teksti, mis on ühes reas
            puhas_tekst = re.sub(r'[^A-Za-z0-9]', '', numbrimärigi_tähed) # regulaar avaldis, et eemaldada mitte regulaarsed tähed
            puhas_tekst=puhas_tekst.replace(" ","").upper() # korrastab numbrimärgi
            if not kas_on_numbrim2rk(puhas_tekst):
                # Kui esimene numbrimärgi kandidaat ei vasta tahetud parameetritel, käib läbi kontuuride järjendi ja vaatab, kas leiab midagi asemele,
                for kontuur in kontuurid:
                    ümbermõõt = cv2.arcLength(kontuur, True)
                    umbkaudne = cv2.approxPolyDP(kontuur, 0.01 * ümbermõõt, True)
                    if 4 <= len(umbkaudne) <= 6:
                        x, y, w, h = cv2.boundingRect(umbkaudne)
                        numbrimärk_pilt = hall[y:y + h, x:x + w]
                        numbrimärk_pilt = cv2.resize(numbrimärk_pilt, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
                        numbrimärk_pilt = cv2.GaussianBlur(numbrimärk_pilt, (5, 5), 0)
                        cv2.imwrite("nmbrim2rk.jpg", numbrimärk_pilt)
                        numbrimärigi_tähed = pytesseract.image_to_string(numbrimärk_pilt, config='--psm 7')
                        puhas_tekst = re.sub(r'[^A-Za-z0-9]', '', numbrimärigi_tähed)
                        puhas_tekst = puhas_tekst.replace(" ","").upper()
                        
                        if kas_on_numbrim2rk(puhas_tekst):
                            break
                if not kas_on_numbrim2rk(puhas_tekst):
                    konsool.insert("end","Ei saanud numbrimärgist aru, proovi teist pilti või kirjuta firma lahtrisse ja analüüsi uuesti")
                    return
            tekstvar.set(f"Numbrimärk on {puhas_tekst}")
            konsool.insert("end","Sisesta liiva kogus")

        else:
            konsool.insert("end","Ei saanud numbrimärgist aru, proovi teist pilti või kirjuta number lahtrisse ja analüüsi uuesti")
    except Exception as e:
        konsool.insert("end","VIGA", e)
        print(e)
    konsool.configure(state="disabled")

def salvesta():
    konsool.configure(state="normal")
    konsool.delete(0.0,"end")
    global excel
    global puhas_tekst
    failinimi=excel
    numbrimärk=puhas_tekst
    firma=""
    if len (firmaa.get())!=0:
        firma=firmaa.get()
    if excel==None or puhas_tekst==None:
        konsool.insert("end","Oled jätnud exceli faili lisamata või ei ole analüüsinud numbrimärki")
        konsool.configure(state="disabled")
        return
    #Avab exceli faili
    df=pd.read_excel(failinimi, sheet_name='Avakuva')
    wb=load_workbook(failinimi)
    
    #Kas nr on failis
    kas_on_ridades=df.isin([numbrimärk]).any(axis=1)
    firmad={}
    täis=False
    if len(firma)==0:
        #Kui on failis, siis mis firma all ning lisab sõnastikku
        if kas_on_ridades.any():
            read=df.index[kas_on_ridades].tolist()
            for reanumber in read:
                firmad[df.loc[reanumber,'Firma']]=reanumber
            firmade_list=list(firmad.keys())
        if len(firmad)!=0:
            if len(firmad)==1:
                firma=firmade_list[0]
            else:
                if len(firmaa.get())==0:
                    konsool.insert("end",f"Numbrimärk {numbrimärk} kuulub firmadele: {' ja '.join(firmad)}\nKummale firmale numbrimärk kuulub? (Kirjuta lahtrisse ja salvesta uuesti)")
                    konsool.configure(state="disabled")
                    return
                firma=firmaa.get()
            firmarida=firmad[firma]
        else:
            if len(firmaa.get())>0:
                firma=firmaa.get()
            else:
                konsool.insert("end",f"Numbrimärki {numbrimärk} ei ole klientide nimekirjas, kui tahad lisada, kirjuta lahtrisse ja salvesta uuesti.")
                konsool.configure(state="disabled")
                return
    try:
        reanr=df.loc[df['Firma'] == firma].index[0]+2
    except:
        konsool.insert("end",f"Sellist firmat ei ole nimekirjas.")
        konsool.configure(state="disabled")
        return
    firmarida=reanr+2
    #Palju krediidilimiiti alles on?
    ws = wb['Avakuva']
    df2=pd.read_excel(failinimi, sheet_name=firma)
    krediidilimiit=ws.cell(row=reanr, column=5).value
    ws=wb[firma]
    tühirida=0
    # Mitu rida
    for rida in ws:
        if any(kast.value is not None for kast in rida):
            tühirida+=1
    # Summa, sest exceli funktsioonid ei uuenda, kui fail kinni
    kokku_liiva_võtnud=float(df2.loc[0:tühirida,'Kogus (t)'].sum())
    try:
        materjali_kogus=float(liivakogus.get())
    except:
        konsool.insert("end",f"Sisesta korrektne liivakogus.")
        konsool.configure(state="disabled")
        return
    if krediidilimiit -kokku_liiva_võtnud<materjali_kogus:
        konsool.insert("end",f"Krediidilimiit on ületatud {kokku_liiva_võtnud+materjali_kogus-krediidilimiit}t võrra, ei saa liiva anda.")
        konsool.configure(state="disabled")
        return
    kuupäev=datetime.now().strftime('%d.%m.%y')

    #Otsib esimese tühi lahtri selle firma alalehel ja kirjutab sinna vajalikud andmed
    ws.cell(column=1, row=tühirida+1, value=numbrimärk)
    ws.cell(column=2, row=tühirida+1, value=kuupäev)
    ws.cell(column=3, row=tühirida+1, value=materjali_kogus)
    ws.cell(column=4, row=tühirida+1, value=f'=C{tühirida+1}*Avakuva!$F${firmarida}')
    wb.save(failinimi)

    #Prindib tulemuse
    konsool.insert("end",f"Numbrimärk {numbrimärk} võttis {materjali_kogus}t liiva {datetime.now().strftime('%d.%m.%y %H:%M')} {firma} nimele\n")
    #Kui täis
    if täis==False:
        konsool.insert("end",f"Krediidilimiidi täitumiseni saab veel võtta {krediidilimiit -(kokku_liiva_võtnud+materjali_kogus)}t liiva.")
    konsool.configure(state="disabled")

excel=None
pilttt=None
img_label=None
puhas_tekst=None
backup=None
windows_path = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
pytesseract.tesseract_cmd = windows_path
windll.shcore.SetProcessDpiAwareness(1)

# Aken
customtkinter.set_appearance_mode("dark")
customtkinter.set_default_color_theme("blue")
app=customtkinter.CTk()
app.title("Numbrimärgilugeja")
app.geometry("500x500")
app.minsize(500, 500)
app.maxsize(500,500)

# Vasak pool
vasak=customtkinter.CTkFrame(master=app,
                  width=240,
                  height=490,
                  corner_radius=10)
vasak.place(relx=0,rely=0.01)
# Kasutamata nupp
# nupp=customtkinter.CTkButton(master=vasak, text="Pildista",command=pildista,width=100)
# nupp.place(relx=0.25,rely=0.1,anchor=tkinter.CENTER)
nupp1=customtkinter.CTkButton(master=vasak, text="Vali pilt",command=valipilt)
nupp1.place(relx=0.5,rely=0.8,anchor=tkinter.CENTER)
nupp2=customtkinter.CTkButton(master=vasak, text="Vali excel",command=valiexcel)
nupp2.place(relx=0.5,rely=0.9,anchor=tkinter.CENTER)
nupp4=customtkinter.CTkButton(master=vasak, text="Analüüsi",command=analuus,width=100)
nupp4.place(relx=0.5,rely=0.1,anchor=tkinter.CENTER)

#Parem pool
tekstvar=tkinter.StringVar()
tekstvar.set("Numbrimärk on")
nrm2rk=customtkinter.CTkEntry(master=app,
                            placeholder_text="Numbrimärk",
                            width=150,
                            height=30,
                            border_width=2,
                            corner_radius=10)
nrm2rk.place(relx=0.65,rely=0.16,anchor=tkinter.CENTER)
number=customtkinter.CTkLabel(master=app,
                                textvariable=tekstvar)
number.place(relx=0.5,rely=0.1,anchor=tkinter.W)
liivakogus=customtkinter.CTkEntry(master=app,
                                  placeholder_text="Sisesta liiva kogus",
                                  width=150,
                                  height=30,
                                  border_width=2,
                                  corner_radius=10)
liivakogus.place(relx=0.65,rely=0.35,anchor=tkinter.CENTER)
firmaa=customtkinter.CTkEntry(master=app,
                            placeholder_text="Muuda firmat",
                            width=150,
                            height=30,
                            border_width=2,
                            corner_radius=10)
firmaa.place(relx=0.65,rely=0.5,anchor=tkinter.CENTER)
liiv=customtkinter.CTkLabel(master=app,
                                text="Liiva kogus",
                                width=80,
                                height=25)
firm=customtkinter.CTkLabel(master=app,
                                text="Muuda vajadusel firma nime",
                                width=120,
                                height=25)
liiv.place(relx=0.5,rely=0.29,anchor=tkinter.W)
firm.place(relx=0.5,rely=0.44,anchor=tkinter.W)
#Nupud
salvestanupp=customtkinter.CTkButton(master=app,text="Salvesta", hover_color="green",width=100,command=salvesta)
katkestanupp=customtkinter.CTkButton(master=app,text="Katkesta", hover_color="red",width=100,command=app.destroy)
salvestanupp.place(relx=0.6,rely=0.65,anchor=tkinter.CENTER)
katkestanupp.place(relx=0.85,rely=0.65,anchor=tkinter.CENTER)

konsool=customtkinter.CTkTextbox(app,width=250,height=150,wrap="word")
konsool.grid(row=0,column=0)
tekst="Siia tulevad teavitused\n"
konsool.insert("0.0",tekst)
konsool.place(x=245,y=350)
konsool.configure(state="disabled")

#Hoiab programmi töös
app.mainloop()